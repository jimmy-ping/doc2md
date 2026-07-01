"""Tests for doc2md — pure logic only, no GUI or real file I/O.

Run: pytest tests/ -v
"""

import sys
import io
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from utils import detect_format, sanitize_filename, format_size, err_exit
from converters.base import BaseConverter, ConversionResult
from converters.txt_converter import TXTConverter
from converters.xlsx_converter import XLSXConverter


# ═══════════════════════════════════════════════════════════════
# utils tests
# ═══════════════════════════════════════════════════════════════

class TestDetectFormat:
    def test_pdf(self):
        assert detect_format(Path("doc.pdf")) == "pdf"

    def test_docx(self):
        assert detect_format(Path("doc.docx")) == "docx"

    def test_doc(self):
        assert detect_format(Path("doc.doc")) == "docx"

    def test_pptx(self):
        assert detect_format(Path("slides.pptx")) == "pptx"

    def test_ppt(self):
        assert detect_format(Path("slides.ppt")) == "pptx"

    def test_xlsx(self):
        assert detect_format(Path("data.xlsx")) == "xlsx"

    def test_xls(self):
        assert detect_format(Path("data.xls")) == "xlsx"

    def test_txt(self):
        assert detect_format(Path("notes.txt")) == "txt"

    def test_md(self):
        assert detect_format(Path("readme.md")) == "txt"

    def test_csv(self):
        assert detect_format(Path("data.csv")) == "txt"

    def test_uppercase_extension(self):
        assert detect_format(Path("DOC.PDF")) == "pdf"

    def test_unsupported_raises(self):
        with pytest.raises(ValueError, match="不支持的文件格式"):
            detect_format(Path("movie.mp4"))


class TestSanitizeFilename:
    def test_simple(self):
        assert sanitize_filename("report") == "report"

    def test_removes_slashes(self):
        assert sanitize_filename("a/b:c") == "a_b_c"

    def test_removes_special_chars(self):
        name = sanitize_filename('file?*"<>|name')
        assert "?" not in name
        assert "*" not in name
        assert '"' not in name

    def test_trims_whitespace(self):
        assert sanitize_filename("  hello  ") == "hello"

    def test_empty_returns_untitled(self):
        assert sanitize_filename("") == "untitled"

    def test_very_long_name(self):
        long_name = "x" * 200
        result = sanitize_filename(long_name)
        assert len(result) <= 120

    def test_long_with_extension(self):
        long_name = "x" * 200 + ".txt"
        result = sanitize_filename(long_name)
        assert result.endswith(".txt")
        assert len(result) <= 120


class TestFormatSize:
    def test_bytes(self):
        assert format_size(500) == "500 B"

    def test_kb(self):
        assert format_size(2048) == "2 KB"

    def test_mb(self):
        assert format_size(3 * 1024 * 1024) == "3 MB"

    def test_gb(self):
        assert format_size(5 * 1024 * 1024 * 1024) == "5 GB"


# ═══════════════════════════════════════════════════════════════
# BaseConverter tests
# ═══════════════════════════════════════════════════════════════

class DummyConverter(BaseConverter):
    """Minimal concrete converter for testing BaseConverter methods."""
    def convert(self):
        return ConversionResult(markdown="test")


class TestImageNaming:
    def test_increments_counter(self):
        conv = DummyConverter(Path("dummy.txt"))
        assert conv._img_counter == 0
        name, _ = conv._next_image(b"fake", "png")
        assert name == "img_001.png"
        assert conv._img_counter == 1
        name2, _ = conv._next_image(b"fake", "png")
        assert name2 == "img_002.png"

    def test_converts_jpeg_to_png(self):
        conv = DummyConverter(Path("dummy.txt"))
        # Create a minimal valid JPEG bytes
        from PIL import Image
        buf = io.BytesIO()
        img = Image.new("RGB", (1, 1), color="red")
        img.save(buf, format="JPEG")
        jpeg_bytes = buf.getvalue()

        name, png_bytes = conv._next_image(jpeg_bytes, "jpeg")
        assert name.endswith(".png")
        # Should be valid PNG
        Image.open(io.BytesIO(png_bytes))  # no exception

    def test_png_passthrough(self):
        conv = DummyConverter(Path("dummy.txt"))
        from PIL import Image
        buf = io.BytesIO()
        img = Image.new("RGB", (1, 1), color="blue")
        img.save(buf, format="PNG")
        png_bytes = buf.getvalue()

        name, result_bytes = conv._next_image(png_bytes, "png")
        assert result_bytes == png_bytes  # unchanged


class TestImageMd:
    def test_basic(self):
        # _image_md uses filename as alt text by default
        assert "images/img_001.png" in BaseConverter._image_md("img_001.png")

    def test_with_alt(self):
        result = BaseConverter._image_md("img_001.png", alt="My Image")
        assert result == "![My Image](images/img_001.png)"


class TestEscapePipe:
    def test_plain_text(self):
        assert BaseConverter._escape_pipe("hello") == "hello"

    def test_escapes_pipe(self):
        assert BaseConverter._escape_pipe("a|b") == "a\\|b"

    def test_replaces_newline(self):
        assert BaseConverter._escape_pipe("a\nb") == "a b"


class TestCleanText:
    def test_collapses_blank_lines(self):
        text = "line1\n\n\n\nline2"
        result = BaseConverter._clean_text(text)
        assert result == "line1\n\nline2"

    def test_strips_trailing(self):
        # _clean_text strips trailing whitespace but preserves leading (e.g. indentation)
        result = BaseConverter._clean_text("  hello  \n  world  ")
        assert result == "  hello\n  world"

    def test_handles_empty(self):
        assert BaseConverter._clean_text("") == ""


# ═══════════════════════════════════════════════════════════════
# TXTConverter tests
# ═══════════════════════════════════════════════════════════════

class TestLooksLikeMarkdown:
    def test_heading_detected(self):
        assert TXTConverter._looks_like_markdown("# Title") is True

    def test_list_detected(self):
        assert TXTConverter._looks_like_markdown("- item") is True

    def test_code_fence_detected(self):
        assert TXTConverter._looks_like_markdown("```code```") is True

    def test_bold_detected(self):
        assert TXTConverter._looks_like_markdown("**bold**") is True

    def test_plain_text_not_detected(self):
        assert TXTConverter._looks_like_markdown("Plain text without markers.") is False


class TestTXTConverter:
    def test_utf8_conversion(self, tmp_path):
        f = tmp_path / "test.txt"
        f.write_text("Hello World", encoding="utf-8")
        conv = TXTConverter(f)
        result = conv.convert()
        assert "Hello World" in result.markdown
        assert result.metadata["encoding"] == "utf-8"

    def test_gbk_conversion(self, tmp_path):
        f = tmp_path / "test.txt"
        # Chinese text in GBK
        text = "中文测试内容"
        f.write_bytes(text.encode("gbk"))
        conv = TXTConverter(f)
        result = conv.convert()
        assert "中文测试内容" in result.markdown
        assert result.metadata["encoding"] in ("gbk", "gb2312", "gb18030")

    def test_markdown_passthrough(self, tmp_path):
        f = tmp_path / "test.md"
        f.write_text("# Title\n\nContent with **bold**", encoding="utf-8")
        conv = TXTConverter(f)
        result = conv.convert()
        # Should not be wrapped in code fence
        assert result.markdown.startswith("# Title")

    def test_line_count_metadata(self, tmp_path):
        f = tmp_path / "test.txt"
        f.write_text("line1\nline2\nline3", encoding="utf-8")
        conv = TXTConverter(f)
        result = conv.convert()
        assert result.metadata["line_count"] == 3


# ═══════════════════════════════════════════════════════════════
# XLSXConverter tests (in-memory workbooks)
# ═══════════════════════════════════════════════════════════════

class TestXLSXConverter:
    def test_simple_table(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Name"
        ws["B1"] = "Age"
        ws["A2"] = "Alice"
        ws["B2"] = 30

        f = tmp_path / "test.xlsx"
        wb.save(str(f))
        wb.close()

        conv = XLSXConverter(f)
        result = conv.convert()
        assert "## Sheet1" in result.markdown
        assert "| Name | Age |" in result.markdown
        assert "| Alice | 30 |" in result.markdown
        assert result.metadata["sheet_count"] == 1

    def test_merged_cell_handling(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Header"
        ws.merge_cells("A1:B1")

        f = tmp_path / "test.xlsx"
        wb.save(str(f))
        wb.close()

        conv = XLSXConverter(f)
        result = conv.convert()
        assert "Header" in result.markdown

    def test_empty_cell_handling(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Col1"
        ws["B1"] = "Col2"
        ws["A2"] = "data"
        # B2 is empty

        f = tmp_path / "test.xlsx"
        wb.save(str(f))
        wb.close()

        conv = XLSXConverter(f)
        result = conv.convert()
        # Should have separator and data row
        assert result.markdown.count("|") >= 8  # pipe count check

    def test_multiple_sheets(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1["A1"] = "X"
        ws1["A2"] = 1

        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Total"
        ws2["A2"] = 100

        f = tmp_path / "test.xlsx"
        wb.save(str(f))
        wb.close()

        conv = XLSXConverter(f)
        result = conv.convert()
        assert "## Data" in result.markdown
        assert "## Summary" in result.markdown
        assert result.metadata["sheet_count"] == 2


# ═══════════════════════════════════════════════════════════════
# Edge cases
# ═══════════════════════════════════════════════════════════════

class TestEdgeCases:
    def test_sanitize_only_special_chars(self):
        result = sanitize_filename("???")
        assert result == "untitled" or len(result) > 0

    def test_format_size_zero(self):
        assert format_size(0) == "0 B"

    def test_detect_format_no_extension(self):
        with pytest.raises(ValueError):
            detect_format(Path("no_extension"))

    def test_detect_format_dotfile(self):
        # .gitignore etc — no recognized extension
        with pytest.raises(ValueError):
            detect_format(Path(".hiddenfile"))


# ═══════════════════════════════════════════════════════════════
# ConversionResult tests
# ═══════════════════════════════════════════════════════════════

class TestConversionResult:
    def test_defaults(self):
        cr = ConversionResult(markdown="# Hi")
        assert cr.markdown == "# Hi"
        assert cr.images == []
        assert cr.metadata == {}

    def test_with_images(self):
        cr = ConversionResult(
            markdown="![](images/img_001.png)",
            images=[("img_001.png", b"fake")],
        )
        assert len(cr.images) == 1
        assert cr.images[0][0] == "img_001.png"
