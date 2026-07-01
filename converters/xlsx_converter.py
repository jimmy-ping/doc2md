"""Excel (.xlsx/.xls) → Markdown table converter using openpyxl."""

from pathlib import Path

import openpyxl

from converters.base import BaseConverter, ConversionResult


class XLSXConverter(BaseConverter):
    """Convert Excel workbooks to Markdown tables."""

    MAX_ROWS = 2000  # Safety limit per sheet
    MAX_COLS = 100

    def convert(self) -> ConversionResult:
        wb = openpyxl.load_workbook(str(self.file_path), data_only=True)

        md_parts: list[str] = []
        images: list[tuple[str, bytes]] = []

        # Workbook title
        # md_parts.append(f"# {self.file_path.stem}\n")

        for ws in wb.worksheets:
            sheet_md, sheet_images = self._convert_sheet(ws)
            if sheet_md.strip():
                md_parts.append(sheet_md)
            images.extend(sheet_images)

        wb.close()

        metadata = {
            "format": "xlsx",
            "sheet_count": len(wb.worksheets),
            "sheets": ", ".join(ws.title for ws in wb.worksheets),
        }

        return ConversionResult(markdown="\n\n".join(md_parts), images=images, metadata=metadata)

    def _convert_sheet(
        self, ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> tuple[str, list[tuple[str, bytes]]]:
        """Convert one worksheet to markdown table."""
        lines: list[str] = []
        images: list[tuple[str, bytes]] = []

        lines.append(f"## {ws.title}\n")

        # Determine data range
        max_row = min(ws.max_row or 0, self.MAX_ROWS)
        max_col = min(ws.max_column or 0, self.MAX_COLS)

        if max_row == 0 or max_col == 0:
            return "", []

        # Build merged cell lookup
        merged_map: dict[tuple[int, int], str] = {}
        for merged_range in ws.merged_cells.ranges:
            val = ws.cell(merged_range.min_row, merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(row, col)] = str(val or "")

        # Collect rows (first row = header)
        headers: list[str] = []
        data_rows: list[list[str]] = []

        is_header_done = False
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=False), 1):
            cells: list[str] = []
            for col_idx, cell in enumerate(row, 1):
                # Check merged cells
                if (row_idx, col_idx) in merged_map:
                    cells.append(merged_map[(row_idx, col_idx)])
                else:
                    val = cell.value
                    cells.append(str(val) if val is not None else "")

            # Skip completely empty rows
            if all(c == "" for c in cells):
                continue

            if not is_header_done:
                headers = cells
                is_header_done = True
            else:
                data_rows.append(cells)

        if not headers:
            return "", []

        # Build pipe table
        # Header row
        escaped_headers = [self._escape_pipe(h) for h in headers]
        lines.append("| " + " | ".join(escaped_headers) + " |")
        # Separator
        lines.append("| " + " | ".join("---" for _ in headers) + " |")
        # Data rows
        for row in data_rows:
            # Pad short rows
            while len(row) < len(headers):
                row.append("")
            escaped = [self._escape_pipe(c) for c in row[:len(headers)]]
            lines.append("| " + " | ".join(escaped) + " |")

        # Extract images from sheet
        if hasattr(ws, "_images"):
            for img in ws._images:
                try:
                    img_data = img._data()
                    filename, png_bytes = self._next_image(img_data, "png")
                    images.append((filename, png_bytes))
                    lines.append("")
                    lines.append(self._image_md(filename))
                except Exception:
                    continue

        return "\n".join(lines), images
